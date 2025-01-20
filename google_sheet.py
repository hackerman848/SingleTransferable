import tools as t
import classes as c
import pandas as pd
import openpyxl as opx
import xlwings as xw
import os
from tkinter import messagebox

#sa = gspread.service_account(filename = "service_account.json")


def load(sheet: str, tab_name: str) -> list[list[str]]:
	# sh = sa.open(sheet)
	sh = pd.read_excel(sheet,sheet_name=0)
	wks = sh.worksheet(tab_name)
	return wks.get_all_values()


def write_results2(vote_inc,sheetname):
	# workbook = opx.load_workbook("template2.xlsx")
	workbook = opx.load_workbook("template.xlsx")
	sheet = workbook.active
	vote: c.Vote = t.deepcopy(vote_inc)
	data = [(len(vote.get_original_ballots()) + vote.get_original_expired()),len(vote.get_original_ballots()),"",vote.get_original_expired(),"",vote.seats]
	print(data)
	write_column(sheet,sheet["AA3"],data)
	sheet["AA11"] = vote.quota
 
	writer = []
	for elected_person in vote.get_all_elected():
		elected_vote_count = vote.get_election_votes(elected_person)
		writer.append(['', '', '', '', '', '', elected_person, '', '', vote.tabulation_rounds[0].get_starting_vote_count()[elected_person], elected_vote_count])
	print(writer)
	# startrow = "D5"
	# endrow = f"D{(len(writer) + 4)}"
	for i in range(len(writer)):
		write_row(sheet,sheet[f"D{i+5}"],writer[i])
 
	
	random_logs = vote.get_all_random_logs()
	
	writer = []
	debuginfo = []
	for i, logs in enumerate(random_logs):
		tab_round = i + 1
		for log in logs:
			log: c.RandLog
			writer.append([ '', '', '', '',f"Round {tab_round}",  '', '', '', '',log.chosen,  '', '', '', '', '', '', '', '', '', '', '', '',"; ".join(log.pool), log.random_for])
			debuginfo.append((tab_round,log.chosen,log.pool,log.random_for))
	print(debuginfo)
	for i in range(len(writer)):
		write_row(sheet,sheet[f"AD{i+5}"],writer[i])
	
	for i, tabulation_round in enumerate(vote.tabulation_rounds):
		writer = []
		for person in tabulation_round.get_all_starting_candidates():
			if person in tabulation_round.elected:
				event = "Elected"
			elif person in tabulation_round.eliminated:
				event = "Eliminated"
			else:
				event = "- -"

			if tabulation_round.outgoing_ballots is not None:
				try:
					turn2 = tabulation_round.outgoing_vote_count[person]
				except KeyError:
					turn2 = "- -"
			else:
				turn2 = "<no transfer>"

			writer.append([ '', '', '', '', '', '', '', '', '', '', '', '', '',person,  '', '', '',tabulation_round.get_starting_vote_count()[person],  '', '', '',turn2, event])
		print("i = "+ str(i))
		print(writer)
		print("\n")
		if i % 3 == 0:
			columns = ("C", "Y")
		elif i % 3 == 1:
			columns = ("AE", "BA")
		else:
			columns = ("BG", "CC")

		row_1 = (i // 3) * 34 + 39
		row_2 = row_1 + 29

		for j in range(len(writer)):
			write_row(sheet,sheet[f"{columns[0]}{row_1+j}"],writer[j])
	
	
	workbook.save("temp.xlsx")
	stitch("temp.xlsx",sheetname)
	os.remove("temp.xlsx")
	print("saved")
	messagebox.showinfo("Information",f"Calculation finished, see {sheetname}")


def find_merged_cell(sheet, cell):
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return merged_range.min_row, merged_range.min_col
    return cell.row, cell.column


def write_row(sheet, start_cell, data):
    start_row, start_col = find_merged_cell(sheet, start_cell)
    for i, value in enumerate(data):
        row, col = start_row, start_col + i
        row, col = find_merged_cell(sheet, sheet.cell(row=row, column=col))
        sheet.cell(row=row, column=col, value=value)

def write_column(sheet, start_cell, data):
    start_row, start_col = find_merged_cell(sheet, start_cell)
    for i, value in enumerate(data):
        row, col = start_row + i, start_col
        row, col = find_merged_cell(sheet, sheet.cell(row=row, column=col))
        sheet.cell(row=row, column=col, value=value)

def stitch(source,destination):
	app = xw.App(visible=False)

	sourcewb = app.books.open(source)
	destinationwb = app.books.open(destination)

	sourcesheet = sourcewb.sheets[0]

	try:
		destinationwb.sheets["Results"].delete()
	except: #already exists
		pass

	sourcesheet.copy(after=destinationwb.sheets[-1])

	destinationwb.sheets[-1].name = "Results"

	destinationwb.save(destination)
	sourcewb.close()
	destinationwb.close()

	app.quit()

